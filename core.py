"""
core.py — Shared attendance-processing logic.

This module is intentionally free of FastAPI/web-framework imports so it can
be used both by the HTTP API (main.py) and directly from the CLI or tests.

Public API
----------
    normalizar_nombre(nombre)           -> str
    extraer_participantes_zoom(html)    -> dict
    procesar_asistencia(in, out, src)   -> list[list[str]]
"""
import csv
import json
import os
import re
import unicodedata
from bs4 import BeautifulSoup
import openpyxl


def normalizar_nombre(nombre: str) -> str:
    """
    Normaliza un nombre para facilitar la comparación:
    - Remueve tildes/acentos (ej. María -> Maria, Pérez -> Perez)
    - Pasa a minúsculas
    - Remueve espacios redundantes
    - Remueve sufijos típicos de Zoom como (yo), (anfitrión), (host), (guest), etc.
    """
    if not nombre:
        return ""
    
    # 1. Normalizar caracteres unicode (quitar diacríticos/tildes)
    nfkd = unicodedata.normalize('NFKD', str(nombre))
    sin_tildes = "".join([c for c in nfkd if not unicodedata.combining(c)])
    
    # 2. Convertir a minúsculas y limpiar espacios
    limpio = re.sub(r'\s+', ' ', sin_tildes).strip().lower()
    
    # 3. Remover sufijos de Zoom si estuvieran dentro del texto
    limpio = re.sub(
        r'\s*\((yo|anfitri[oó]n|coanfitri[oó]n|host|co-host|me|guest|invitado)[^)]*\)',
        '',
        limpio,
        flags=re.IGNORECASE
    )
    
    return limpio.strip()


def extraer_participantes_zoom(html_content: str) -> dict:
    """
    Extrae los participantes y el estado de sus cámaras a partir del HTML de Zoom.
    Soporta la estructura virtualizada de Zoom Web (formato-ejemplo.html).
    
    Retorna un diccionario con la estructura:
    {
        "nombre_normalizado": {
            "nombre_original": str,
            "camera_on": bool,
            "aria_label": str
        }
    }
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    participantes = {}
    
    # 1. Buscar elementos de participante:
    # Zoom web utiliza elementos con clase 'participants-li' o IDs con 'participants-list-'
    items = soup.find_all(
        lambda tag: tag.has_attr('class') and 'participants-li' in tag.get('class', [])
    )
    
    if not items:
        # Fallback: buscar por ID
        items = soup.find_all(
            lambda tag: tag.has_attr('id') and str(tag.get('id', '')).startswith('participants-list-')
        )
        
    if not items:
        # Fallback general: buscar contenedores con layout de participante
        items = soup.find_all(
            class_=lambda c: c and any(cls in ['participants-item__item-layout', 'participants-item-position'] for cls in (c if isinstance(c, list) else c.split()))
        )

    for item in items:
        # 2. Extraer nombre desde el span con clase 'participants-item__display-name'
        name_el = item.find(
            class_=lambda c: c and 'participants-item__display-name' in (c if isinstance(c, list) else c.split())
        )
        
        nombre_original = ""
        if name_el:
            nombre_original = name_el.get_text(strip=True)
            
        aria_label = item.get('aria-label', '')
        
        # Si no encontramos el span de nombre, intentar extraerlo del aria-label
        if not nombre_original and aria_label:
            nombre_part = aria_label.split(',')[0].strip()
            nombre_original = re.sub(r'\s*\([^)]*\)', '', nombre_part).strip()
            
        if not nombre_original:
            continue
            
        # 3. Detectar estado de la cámara
        # A. Revisar íconos SVG con clase 'video-off'
        has_video_off_svg = bool(
            item.find('svg', class_=lambda c: c and any('video-off' in cls for cls in (c if isinstance(c, list) else c.split())))
        )
        
        # B. Revisar si en aria-label indica video off o video on
        aria_lower = aria_label.lower()
        has_video_off_aria = any(
            phrase in aria_lower for phrase in ["video off", "video apagado", "camara apagada", "cámara apagada", "video-off"]
        )
        has_video_on_aria = any(
            phrase in aria_lower for phrase in ["video on", "video encendido", "camara encendida", "cámara encendida", "video-on"]
        )
        has_video_on_svg = bool(
            item.find('svg', class_=lambda c: c and any('video-on' in cls for cls in (c if isinstance(c, list) else c.split())))
        )
        
        if has_video_off_svg or has_video_off_aria:
            camera_on = False
        elif has_video_on_svg or has_video_on_aria:
            camera_on = True
        else:
            # En Zoom, cuando la cámara está activa usualmente no muestra el ícono tachado de video-off
            camera_on = True
            
        clave_normalizada = normalizar_nombre(nombre_original)
        participantes[clave_normalizada] = {
            "nombre_original": nombre_original,
            "camera_on": camera_on,
            "aria_label": aria_label
        }
        
    return participantes


def procesar_asistencia(input_csv: str, output_csv: str, zoom_source) -> list:
    """
    Compara la lista de alumnos del CSV con los participantes de Zoom y
    genera un reporte CSV con el estado de asistencia y cámara.

    Parameters
    ----------
    input_csv : str
        Ruta al archivo CSV de alumnos.  Se detecta automáticamente el encoding
        (utf-8-sig, utf-8, latin-1, cp1252).
    output_csv : str
        Ruta donde se escribirá el CSV de salida (siempre en utf-8-sig / BOM,
        para compatibilidad con Excel).
    zoom_source : str | dict
        Puede ser cualquiera de:
        - dict  → {nombre: {"camera_on": bool, ...}}
        - str   → ruta a un archivo HTML/JSON en disco
        - str   → HTML crudo (contiene "<" y palabras clave de Zoom)
        - str   → JSON crudo parseable con json.loads()

    Returns
    -------
    list[list[str]]
        Lista de filas incluyendo el encabezado::

            [["Nombre", "Asistencia", "Cámara"],
             ["María López", "Presente", "Encendida"],
             ["Carlos Diaz", "Ausente", "Apagada"],
             ...]
    """
    # 1. Resolver y normalizar zoom_data
    zoom_dict_normalizado = {}
    
    if isinstance(zoom_source, dict):
        for k, v in zoom_source.items():
            norm_k = normalizar_nombre(k)
            if isinstance(v, dict):
                zoom_dict_normalizado[norm_k] = v
            else:
                zoom_dict_normalizado[norm_k] = {"camera_on": bool(v)}
                
    elif isinstance(zoom_source, str):
        fuente = zoom_source.strip()
        # Verificar si es una ruta a un archivo existente
        if os.path.isfile(fuente):
            with open(fuente, 'r', encoding='utf-8', errors='ignore') as f:
                contenido = f.read()
            if "<" in contenido and ("participants" in contenido or "html" in contenido or "div" in contenido):
                zoom_dict_normalizado = extraer_participantes_zoom(contenido)
            else:
                try:
                    data = json.loads(contenido)
                    zoom_dict_normalizado = {normalizar_nombre(k): v if isinstance(v, dict) else {"camera_on": bool(v)} for k, v in data.items()}
                except json.JSONDecodeError:
                    zoom_dict_normalizado = extraer_participantes_zoom(contenido)
        # Verificar si es HTML directo
        elif "<" in fuente and ("participants" in fuente or "div" in fuente or "svg" in fuente):
            zoom_dict_normalizado = extraer_participantes_zoom(fuente)
        # Verificar si es JSON
        else:
            try:
                data = json.loads(fuente)
                if isinstance(data, dict):
                    zoom_dict_normalizado = {normalizar_nombre(k): v if isinstance(v, dict) else {"camera_on": bool(v)} for k, v in data.items()}
            except json.JSONDecodeError:
                # Si no es JSON válido, intentar extraer como HTML
                zoom_dict_normalizado = extraer_participantes_zoom(fuente)

    # 2. Lectura inteligente del archivo CSV con soporte para múltiples encodings
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
    infile = None
    reader = None
    header = None
    
    for enc in encodings:
        try:
            f = open(input_csv, mode='r', encoding=enc)
            test_reader = csv.reader(f)
            header = next(test_reader, None)
            infile = f
            reader = test_reader
            break
        except (UnicodeDecodeError, Exception):
            if f:
                f.close()
            continue
            
    if infile is None:
        raise ValueError(f"No se pudo leer el archivo CSV {input_csv} con codificaciones estándar.")

    # 3. Procesar asistencia de cada alumno
    resultados = []
    col_nombre = header[0].strip() if (header and len(header) > 0 and header[0].strip()) else "Nombre"
    resultados.append([col_nombre, "Asistencia", "Cámara"])
    
    for row in reader:
        if not row or not any(row):
            continue
            
        nombre_alumno_original = row[0].strip()
        if not nombre_alumno_original:
            continue
            
        clave_busqueda = normalizar_nombre(nombre_alumno_original)
        info_zoom = zoom_dict_normalizado.get(clave_busqueda)
        
        if info_zoom:
            asistencia = "Presente"
            camara = "Encendida" if info_zoom.get("camera_on") else "Apagada"
        else:
            asistencia = "Ausente"
            camara = "Apagada"
            
        resultados.append([nombre_alumno_original, asistencia, camara])
        
    infile.close()

    # 4. Generar el archivo CSV de salida con utf-8-sig (compatible con Excel)
    with open(output_csv, mode='w', encoding='utf-8-sig', newline='') as outfile:
        writer = csv.writer(outfile)
        writer.writerows(resultados)
        
    return resultados


def marcar_asistencia_excel(xlsx_path: str, zoom_source, col_offset: int = 1) -> dict:
    """
    Lee un archivo Excel (.xlsx) con una lista de alumnos y escribe el estado de
    asistencia en la columna inmediatamente a la derecha del nombre de cada alumno.

    Reglas:
    - Si el alumno aparece en Zoom → escribe "Sí" en la celda de la derecha.
    - Si el alumno no aparece en Zoom → escribe "No" en la celda de la derecha.
    - Si un nombre de Zoom no está en el Excel → no se hace nada (puede estar en
      otra planilla).

    Parameters
    ----------
    xlsx_path : str
        Ruta al archivo Excel (.xlsx) con la lista de alumnos.
        El archivo se modifica in-place y también se guarda en ``xlsx_path``.
    zoom_source : str | dict
        Mismas opciones que ``procesar_asistencia``: dict normalizado, HTML crudo,
        JSON crudo, o ruta a un archivo HTML/JSON.
    col_offset : int, optional
        Cuántas columnas a la derecha del nombre se escribe la asistencia.
        Por defecto 1 (la columna inmediatamente a la derecha).

    Returns
    -------
    dict
        Resumen con claves ``presentes``, ``ausentes``, ``total`` y ``filas``
        (lista de tuplas (nombre, estado) procesadas).
    """
    # 1. Resolver zoom_source en el mismo formato que procesar_asistencia
    zoom_dict_normalizado = {}

    if isinstance(zoom_source, dict):
        for k, v in zoom_source.items():
            norm_k = normalizar_nombre(k)
            if isinstance(v, dict):
                zoom_dict_normalizado[norm_k] = v
            else:
                zoom_dict_normalizado[norm_k] = {"camera_on": bool(v)}

    elif isinstance(zoom_source, str):
        fuente = zoom_source.strip()
        if os.path.isfile(fuente):
            with open(fuente, 'r', encoding='utf-8', errors='ignore') as f:
                contenido = f.read()
            if "<" in contenido and ("participants" in contenido or "html" in contenido or "div" in contenido):
                zoom_dict_normalizado = extraer_participantes_zoom(contenido)
            else:
                try:
                    data = json.loads(contenido)
                    zoom_dict_normalizado = {
                        normalizar_nombre(k): v if isinstance(v, dict) else {"camera_on": bool(v)}
                        for k, v in data.items()
                    }
                except json.JSONDecodeError:
                    zoom_dict_normalizado = extraer_participantes_zoom(contenido)
        elif "<" in fuente and ("participants" in fuente or "div" in fuente or "svg" in fuente):
            zoom_dict_normalizado = extraer_participantes_zoom(fuente)
        else:
            try:
                data = json.loads(fuente)
                if isinstance(data, dict):
                    zoom_dict_normalizado = {
                        normalizar_nombre(k): v if isinstance(v, dict) else {"camera_on": bool(v)}
                        for k, v in data.items()
                    }
            except json.JSONDecodeError:
                zoom_dict_normalizado = extraer_participantes_zoom(fuente)

    # 2. Abrir el libro Excel
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    presentes = 0
    ausentes = 0
    filas_procesadas = []

    # 3. Iterar sobre todas las filas buscando celdas no vacías en la primera columna
    for row in ws.iter_rows():
        # Buscar la primera celda no vacía de la fila como candidata a nombre de alumno
        primera_celda = row[0] if row else None
        if primera_celda is None:
            continue

        valor = primera_celda.value
        if valor is None:
            continue

        nombre_str = str(valor).strip()
        if not nombre_str:
            continue

        clave_busqueda = normalizar_nombre(nombre_str)
        info_zoom = zoom_dict_normalizado.get(clave_busqueda)

        if info_zoom is not None:
            # Alumno presente en Zoom
            estado = "Sí"
            presentes += 1
        else:
            # Alumno en la lista pero ausente en Zoom → marcar No
            estado = "No"
            ausentes += 1

        # Escribir en la columna a la derecha del nombre
        col_destino = primera_celda.column + col_offset
        ws.cell(row=primera_celda.row, column=col_destino, value=estado)
        filas_procesadas.append((nombre_str, estado))

    # 4. Guardar el libro modificado
    wb.save(xlsx_path)

    return {
        "presentes": presentes,
        "ausentes": ausentes,
        "total": presentes + ausentes,
        "filas": filas_procesadas,
    }